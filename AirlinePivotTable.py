import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

airline_xl = pd.read_excel('Air_Traffic_Passenger_Statistics.xlsx')
airline_xl[['Published Airline', 'Activity Type Code', 'Adjusted Passenger Count']]

#Creating pivot table with index as Published airlines and columns as Activity type, using "Sum" as the aggregator on the Passenger Count
pivot_data = airline_xl.pivot_table(index='Published Airline',columns='Activity Type Code',values='Adjusted Passenger Count',aggfunc='sum').round(0)


#Saving the output to an excel file
pivot_data.to_excel('airline_pivot.xlsx',sheet_name='Report',startrow=4)

wb = load_workbook('airline_pivot.xlsx')
sheet = wb['Report']

# cell references (original spreadsheet) 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#Barchart
barchart = BarChart()

#locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) #not including headers


# adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#location chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Passenger Count By Activity Type Code'
barchart.style = 5 #choose the chart style
wb.save('airline_pivot.xlsx')
